"""Audit original uploads through parsing, chunking, and an existing RAG run.

This command is deliberately read-only and does not call paid providers.  It
distinguishes file readability from parsing accuracy, and labels parser
comparisons as proxies unless a separately curated ground truth is supplied.
Private reports should remain under the ignored ``backend/.private-benchmarks`` path.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import sqlite3
import subprocess
from collections import Counter, defaultdict
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from .production_holdout_benchmark import _corpus_sha256, _load_corpus, _sha256

BACKEND_DIR = Path(__file__).resolve().parent.parent
DEFAULT_SOURCE_DB = BACKEND_DIR.parent / "data" / "meetings.db"
DEFAULT_UPLOADS_DIR = BACKEND_DIR.parent / "data" / "uploads"
_TOKEN_RE = re.compile(r"[a-z0-9]+|[\u3400-\u9fff]", re.IGNORECASE)


def _tokens(text: str) -> list[str]:
    """Return format-insensitive Latin tokens and individual CJK characters."""
    return _TOKEN_RE.findall(text.casefold())


def _overlap(reference: str, observed: str) -> dict[str, float | int]:
    """Multiset token agreement; robust to Markdown and layout-only changes."""
    reference_tokens = _tokens(reference)
    observed_tokens = _tokens(observed)
    common = sum((Counter(reference_tokens) & Counter(observed_tokens)).values())
    recall = common / len(reference_tokens) if reference_tokens else 0.0
    precision = common / len(observed_tokens) if observed_tokens else 0.0
    f1 = 2 * recall * precision / (recall + precision) if recall + precision else 0.0
    return {
        "reference_tokens": len(reference_tokens),
        "observed_tokens": len(observed_tokens),
        "matching_tokens": common,
        "token_recall": recall,
        "token_precision": precision,
        "token_f1": f1,
    }


def _normalized(text: str) -> str:
    return " ".join(text.casefold().split())


def _resolve_upload_path(stored_path: str, uploads_dir: Path) -> Path | None:
    """Resolve a DB upload path without allowing the audit to escape its root."""
    root = uploads_dir.resolve()
    candidates = [Path(stored_path), root / Path(stored_path).name]
    for candidate in candidates:
        try:
            resolved = candidate.resolve()
        except OSError:
            continue
        if resolved.is_relative_to(root) and resolved.is_file():
            return resolved
    return None


def _ffprobe(path: Path) -> dict[str, Any]:
    completed = subprocess.run(
        [
            "ffprobe",
            "-v",
            "error",
            "-show_entries",
            "format=duration:stream=codec_type,codec_name",
            "-of",
            "json",
            str(path),
        ],
        check=True,
        capture_output=True,
        text=True,
        timeout=60,
    )
    data = json.loads(completed.stdout)
    streams = data.get("streams") or []
    if not streams:
        raise ValueError("ffprobe found no media streams")
    duration_raw = (data.get("format") or {}).get("duration")
    return {
        "duration_seconds": float(duration_raw) if duration_raw is not None else None,
        "streams": [
            {"type": stream.get("codec_type"), "codec": stream.get("codec_name")}
            for stream in streams
        ],
    }


def _probe_original(path: Path, file_type: str) -> dict[str, Any]:
    details: dict[str, Any] = {"bytes": path.stat().st_size}
    if not details["bytes"]:
        raise ValueError("file is empty")
    kind = file_type.casefold()
    suffix = path.suffix.casefold()
    if kind == "pdf" or suffix == ".pdf":
        import pymupdf as fitz

        with fitz.open(path) as document:
            if document.page_count < 1:
                raise ValueError("PDF has no pages")
            details["pages"] = document.page_count
    elif kind in {"xls", "xlsx"} or suffix in {".xls", ".xlsx"}:
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            details["worksheets"] = len(workbook.sheetnames)
        finally:
            workbook.close()
    elif kind == "image" or suffix in {".png", ".jpg", ".jpeg", ".webp", ".tiff"}:
        from PIL import Image

        with Image.open(path) as image:
            details.update({"width": image.width, "height": image.height, "format": image.format})
            image.verify()
    elif kind in {"audio", "video"}:
        details.update(_ffprobe(path))
    else:
        with path.open("rb") as handle:
            handle.read(1)
    return details


def _xlsx_reference(path: Path) -> str:
    from openpyxl import load_workbook

    rows: list[str] = []
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        for worksheet in workbook.worksheets:
            for row in worksheet.iter_rows(values_only=True):
                values = [str(value) for value in row if value is not None and str(value).strip()]
                if values:
                    rows.append("\t".join(values))
    finally:
        workbook.close()
    return "\n".join(rows)


def _pdf_reference(path: Path) -> str:
    completed = subprocess.run(
        ["pdftotext", "-layout", str(path), "-"],
        check=True,
        capture_output=True,
        text=True,
        errors="replace",
        timeout=120,
    )
    return completed.stdout


def _image_reference(path: Path) -> str:
    completed = subprocess.run(
        ["tesseract", str(path), "stdout", "-l", "eng"],
        check=True,
        capture_output=True,
        text=True,
        errors="replace",
        timeout=120,
    )
    return completed.stdout


def _reference_text(path: Path, file_type: str) -> tuple[str | None, str | None, str | None]:
    """Return reference text, evidence tier, or an explicit skip reason."""
    kind = file_type.casefold()
    try:
        if kind == "pdf":
            text = _pdf_reference(path)
            method = "cross_parser_pdf_text_layer"
        elif kind in {"xls", "xlsx"}:
            text = _xlsx_reference(path)
            method = "structured_cell_projection"
        elif kind == "image":
            text = _image_reference(path)
            method = "independent_local_ocr"
        elif kind in {"txt", "md", "text"}:
            text = path.read_text(encoding="utf-8", errors="replace")
            method = "direct_text_source"
        else:
            return None, None, "independent_reference_unavailable_for_media_transcript"
    except (OSError, subprocess.SubprocessError, ValueError) as exc:
        return None, None, f"reference_extraction_failed:{type(exc).__name__}"
    if len(_tokens(text)) < 10:
        return None, None, "reference_text_too_sparse"
    return text, method, None


def _load_rows(db_path: Path) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    uri = f"file:{db_path.resolve()}?mode=ro"
    with sqlite3.connect(uri, uri=True) as connection:
        connection.row_factory = sqlite3.Row
        files = [
            dict(row)
            for row in connection.execute(
                "SELECT id, meeting_id, file_name, file_type, file_path, content_hash, transcript, "
                "page_count, duration_seconds FROM meeting_files WHERE status='ready' ORDER BY id"
            )
        ]
        chunks: list[dict[str, Any]] = []
        for row in connection.execute(
            "SELECT chunk_id, meeting_id, content, metadata FROM bm25_index ORDER BY id"
        ):
            metadata = json.loads(row["metadata"] or "{}")
            chunks.append(
                {
                    "chunk_id": str(row["chunk_id"]),
                    "meeting_id": row["meeting_id"],
                    "content": str(row["content"] or ""),
                    "metadata": metadata,
                    "file_id": metadata.get("file_id"),
                }
            )
    return files, chunks


def _load_holdout(path: Path) -> tuple[dict[str, Any], dict[int, list[str]]]:
    holdout = json.loads(path.read_text(encoding="utf-8"))
    quotes_by_file: dict[int, list[str]] = defaultdict(list)
    for case in holdout.get("cases", []):
        file_ids = [int(value) for value in case.get("expected_file_ids", [])]
        for file_id in file_ids:
            quotes_by_file[file_id].extend(str(value) for value in case.get("evidence_quotes", []))
    return holdout, quotes_by_file


def _mean(rows: list[dict[str, Any]], key: str) -> float | None:
    values = [float(row[key]) for row in rows if row.get(key) is not None]
    return sum(values) / len(values) if values else None


def _markdown(payload: dict[str, Any]) -> str:
    original = payload["stages"]["original_uploads"]
    parsing = payload["stages"]["parsing"]
    chunking = payload["stages"]["chunking"]
    rag = payload["stages"]["rag_quality"]
    parse_recall = parsing["mean_token_recall"]
    parse_recall_display = "—" if parse_recall is None else f"{parse_recall:.4f}"
    chunk_recall = chunking["mean_transcript_token_recall"]
    chunk_recall_display = "—" if chunk_recall is None else f"{chunk_recall:.4f}"
    correctness = rag["stats"].get("correctness")
    correctness_display = "—" if correctness is None else f"{correctness:.4f}"
    lines = [
        "# Production upload-to-RAG benchmark",
        "",
        f"Generated: `{payload['timestamp']}`",
        "",
        "| Stage | Status | Coverage / primary result |",
        "|---|---:|---:|",
        (
            f"| Original uploads | {original['status']} | {original['readable_files']}/"
            f"{original['total_files']} readable |"
        ),
        (
            f"| Parsing proxy | {parsing['status']} | {parsing['evaluated_files']}/"
            f"{parsing['total_files']} evaluated; recall {parse_recall_display} |"
        ),
        (
            f"| Chunk integrity | {chunking['status']} | transcript recall "
            f"{chunk_recall_display}; evidence "
            f"{chunking['evidence_quotes_survived']}/{chunking['evidence_quotes_total']} |"
        ),
        (
            f"| RAG quality | {rag['status']} | {rag['observed_cases']} cases; correctness "
            f"{correctness_display} |"
        ),
        "",
        "## Interpretation",
        "",
        (
            "Parsing values are token-agreement proxies against a second extraction path, "
            "not human gold accuracy. Audio/video rows are skipped unless an independent "
            "transcript exists. No score is synthesized for skipped rows."
        ),
        "",
        "## Per-file lineage",
        "",
        "| File | Type | Raw | Parse | Parse recall | Chunks | Chunk recall | Holdout evidence |",
        "|---|---|---:|---:|---:|---:|---:|---:|",
    ]
    for row in payload["files"]:
        parse_value = row["parsing"].get("token_recall")
        parse_display = "—" if parse_value is None else f"{parse_value:.4f}"
        chunk_value = row["chunking"].get("transcript_token_recall")
        chunk_display = "—" if chunk_value is None else f"{chunk_value:.4f}"
        evidence = row["chunking"]["evidence_quotes"]
        evidence_display = (
            "—" if evidence["total"] == 0 else f"{evidence['survived']}/{evidence['total']}"
        )
        lines.append(
            f"| {row['file_name']} | {row['file_type']} | {row['original']['status']} | "
            f"{row['parsing']['status']} | {parse_display} | {row['chunking']['chunk_count']} | "
            f"{chunk_display} | {evidence_display} |"
        )
    lines.extend(["", "## Limitations", ""])
    lines.extend(f"- `{item}`" for item in payload["evidence_quality"]["limitations"])
    lines.append("")
    return "\n".join(lines)


def audit(args: argparse.Namespace) -> Path:
    source_db = args.source_db.resolve()
    uploads_dir = args.uploads_dir.resolve()
    holdout_path = args.holdout.resolve()
    rag_result_path = args.rag_result.resolve()
    output = args.output.resolve()
    output.parent.mkdir(parents=True, exist_ok=True)

    files, all_chunks = _load_rows(source_db)
    corpus_chunks, _ = _load_corpus(source_db)
    current_corpus_sha256 = _corpus_sha256(corpus_chunks)
    _holdout, quotes_by_file = _load_holdout(holdout_path)
    rag_result = json.loads(rag_result_path.read_text(encoding="utf-8"))
    chunks_by_file: dict[int, list[dict[str, Any]]] = defaultdict(list)
    for chunk in all_chunks:
        try:
            file_id = int(chunk["file_id"])
        except (TypeError, ValueError):
            continue
        chunks_by_file[file_id].append(chunk)

    file_rows: list[dict[str, Any]] = []
    parse_rows: list[dict[str, Any]] = []
    total_empty_chunks = 0
    total_chunks = 0
    evidence_total = 0
    evidence_survived = 0
    unique_chunk_ids: set[str] = set()
    duplicate_chunk_ids = 0
    missing_metadata = 0
    content_hash_mismatches = 0

    for file_row in files:
        file_id = int(file_row["id"])
        resolved = _resolve_upload_path(str(file_row["file_path"]), uploads_dir)
        original: dict[str, Any]
        reference: str | None = None
        method: str | None = None
        skip_reason: str | None = None
        if resolved is None:
            original = {"status": "fail", "reason": "missing_or_outside_upload_root"}
            skip_reason = "original_unreadable"
        else:
            try:
                details = _probe_original(resolved, str(file_row["file_type"]))
                current_sha256 = _sha256(resolved)
                recorded_hash = str(file_row.get("content_hash") or "")
                hash_matches = bool(recorded_hash) and current_sha256 == recorded_hash
                if not hash_matches:
                    content_hash_mismatches += 1
                original = {
                    "status": "pass",
                    "relative_path": str(resolved.relative_to(uploads_dir)),
                    "sha256": current_sha256,
                    "recorded_content_hash": recorded_hash or None,
                    "recorded_content_hash_matches": hash_matches,
                    **details,
                }
                reference, method, skip_reason = _reference_text(
                    resolved, str(file_row["file_type"])
                )
            except Exception as exc:  # corrupt third-party files can fail in many libraries
                original = {"status": "fail", "reason": f"probe_failed:{type(exc).__name__}"}
                skip_reason = "original_unreadable"

        transcript = str(file_row.get("transcript") or "")
        if reference is None:
            parsing: dict[str, Any] = {
                "status": "not_evaluated",
                "reference_method": None,
                "reason": skip_reason,
                "token_recall": None,
                "token_precision": None,
                "token_f1": None,
            }
        else:
            agreement = _overlap(reference, transcript)
            parsing = {
                "status": "evaluated_proxy",
                "reference_method": method,
                **agreement,
            }
            parse_rows.append(parsing)

        file_chunks = chunks_by_file.get(file_id, [])
        joined_chunks = "\n".join(chunk["content"] for chunk in file_chunks)
        chunk_agreement = _overlap(transcript, joined_chunks) if transcript else None
        normalized_hashes: list[str] = []
        file_empty = 0
        for chunk in file_chunks:
            total_chunks += 1
            chunk_id = chunk["chunk_id"]
            if chunk_id in unique_chunk_ids:
                duplicate_chunk_ids += 1
            unique_chunk_ids.add(chunk_id)
            if not chunk["content"].strip():
                file_empty += 1
                total_empty_chunks += 1
            metadata = chunk["metadata"]
            if any(metadata.get(key) is None for key in ("file_id", "meeting_id", "chunk_index")):
                missing_metadata += 1
            normalized_hashes.append(
                hashlib.sha256(_normalized(chunk["content"]).encode()).hexdigest()
            )

        quotes = quotes_by_file.get(file_id, [])
        survived = sum(_normalized(quote) in _normalized(joined_chunks) for quote in quotes)
        evidence_total += len(quotes)
        evidence_survived += survived
        chunking = {
            "chunk_count": len(file_chunks),
            "empty_chunks": file_empty,
            "duplicate_content_chunks": len(normalized_hashes) - len(set(normalized_hashes)),
            "transcript_token_recall": (
                chunk_agreement["token_recall"] if chunk_agreement is not None else None
            ),
            "transcript_token_precision": (
                chunk_agreement["token_precision"] if chunk_agreement is not None else None
            ),
            "transcript_token_f1": chunk_agreement["token_f1"] if chunk_agreement else None,
            "evidence_quotes": {"total": len(quotes), "survived": survived},
        }
        file_rows.append(
            {
                "file_id": file_id,
                "meeting_id": int(file_row["meeting_id"]),
                "file_name": str(file_row["file_name"]),
                "file_type": str(file_row["file_type"]),
                "original": original,
                "parsing": parsing,
                "chunking": chunking,
                "lineage": {
                    "transcript_sha256": hashlib.sha256(transcript.encode()).hexdigest(),
                    "indexed_content_sha256": hashlib.sha256(joined_chunks.encode()).hexdigest(),
                },
            }
        )

    readable_files = sum(row["original"]["status"] == "pass" for row in file_rows)
    parsing_skipped = len(files) - len(parse_rows)
    chunk_recall_rows = [
        {"value": row["chunking"]["transcript_token_recall"]}
        for row in file_rows
        if row["chunking"]["transcript_token_recall"] is not None
    ]
    mean_chunk_recall = _mean(chunk_recall_rows, "value")
    chunk_valid = bool(
        total_chunks
        and not total_empty_chunks
        and not duplicate_chunk_ids
        and not missing_metadata
        and (mean_chunk_recall or 0.0) >= args.minimum_chunk_recall
        and evidence_total > 0
        and evidence_survived == evidence_total
    )
    rag_consistent = bool(
        rag_result.get("valid")
        and rag_result.get("source_corpus_sha256") == current_corpus_sha256
        and rag_result.get("holdout_sha256") == _sha256(holdout_path)
    )
    original_valid = readable_files == len(files) and content_hash_mismatches == 0
    parsing_status = "partial" if parsing_skipped else "pass"
    limitations = [
        "parse_scores_are_cross_parser_proxies_not_human_gold_accuracy",
        (
            "parser_reference_coverage_is_partial"
            if parsing_skipped
            else "parser_references_not_human_gold"
        ),
    ]
    limitations.extend(
        str(value) for value in rag_result.get("evidence_quality", {}).get("limitations", [])
    )
    limitations = list(dict.fromkeys(limitations))
    valid = original_valid and bool(parse_rows) and chunk_valid and rag_consistent
    payload: dict[str, Any] = {
        "schema_version": 1,
        "command": "production-upload-to-rag-audit",
        "timestamp": datetime.now(UTC).isoformat(),
        "paid_run": False,
        "valid": valid,
        "release_ready": valid and not limitations,
        "source_db": str(source_db),
        "source_corpus_sha256": current_corpus_sha256,
        "holdout": {"path": str(holdout_path), "sha256": _sha256(holdout_path)},
        "rag_result": {"path": str(rag_result_path), "sha256": _sha256(rag_result_path)},
        "stages": {
            "original_uploads": {
                "status": "pass" if original_valid else "fail",
                "total_files": len(files),
                "readable_files": readable_files,
                "readability_rate": readable_files / len(files) if files else 0.0,
                "recorded_content_hash_mismatches": content_hash_mismatches,
            },
            "parsing": {
                "status": parsing_status,
                "total_files": len(files),
                "evaluated_files": len(parse_rows),
                "skipped_files": parsing_skipped,
                "reference_coverage": len(parse_rows) / len(files) if files else 0.0,
                "mean_token_recall": _mean(parse_rows, "token_recall"),
                "mean_token_precision": _mean(parse_rows, "token_precision"),
                "mean_token_f1": _mean(parse_rows, "token_f1"),
                "metric_semantics": "format-insensitive token agreement proxy",
            },
            "chunking": {
                "status": "pass" if chunk_valid else "fail",
                "total_chunks": total_chunks,
                "files_with_chunks": sum(row["chunking"]["chunk_count"] > 0 for row in file_rows),
                "empty_chunks": total_empty_chunks,
                "duplicate_chunk_ids": duplicate_chunk_ids,
                "chunks_missing_required_metadata": missing_metadata,
                "mean_transcript_token_recall": mean_chunk_recall,
                "minimum_required_mean_recall": args.minimum_chunk_recall,
                "evidence_quotes_total": evidence_total,
                "evidence_quotes_survived": evidence_survived,
            },
            "rag_quality": {
                "status": "pass" if rag_consistent else "fail",
                "artifact_valid": bool(rag_result.get("valid")),
                "artifact_consistent_with_current_corpus": rag_consistent,
                "observed_cases": rag_result.get("evidence_quality", {}).get("observed_cases"),
                "judge_repeats": rag_result.get("judge_repeats"),
                "models": {
                    "system": rag_result.get("system_model"),
                    "judge": rag_result.get("judge_model"),
                    "reranker": rag_result.get("reranker_model"),
                },
                "stats": rag_result.get("stats", {}),
            },
        },
        "evidence_quality": {
            "grade": "candidate_requires_review" if valid else "invalid",
            "limitations": limitations,
        },
        "files": file_rows,
    }
    output.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    markdown_path = output.with_suffix(".md")
    markdown_path.write_text(_markdown(payload), encoding="utf-8")
    return output


def _parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--source-db", type=Path, default=DEFAULT_SOURCE_DB)
    parser.add_argument("--uploads-dir", type=Path, default=DEFAULT_UPLOADS_DIR)
    parser.add_argument("--holdout", type=Path, required=True)
    parser.add_argument("--rag-result", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--minimum-chunk-recall", type=float, default=0.98)
    return parser


def main() -> None:
    output = audit(_parser().parse_args())
    print(f"wrote {output}")
    print(f"wrote {output.with_suffix('.md')}")


if __name__ == "__main__":
    main()
